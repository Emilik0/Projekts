
from openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.styles import NamedStyle

wb = load_workbook("Workout.xlsx")
current_date = datetime.now().strftime('%Y-%m-%d')
ws = wb.active


total_time=7 # (min) atpūta starp vingrinājumiem

for i in range(2,ws.max_row+1):
    reps = ws['B' + str(i)].value
    sets = ws['C' + str(i)].value
    spent_time = ws['D'+str(i)].value
    if reps is not None and sets is not None:
        time = float((reps*sets*0.06)+0.5) 
    elif spent_time is not None:
        time = float(spent_time)
    else:
        continue
    total_time += time
total_time = round(total_time)
ws['F2'] = str(total_time) + ' min'

count=0
for i in range(2,ws.max_row+1):
    reps = ws['B' + str(i)].value
    sets = ws['C' + str(i)].value
    spent_time = ws['D'+str(i)].value
    if (reps is not None and sets is not None) or spent_time is not None:
        count+=1
efficiency = round((count/(ws.max_row-1))*100)
ws['E2'] = str(efficiency) + '%'

last_sheet_name = wb.sheetnames[-1]
new_ws = wb.copy_worksheet(wb[last_sheet_name])
new_ws.title = current_date

for row in new_ws.iter_rows(min_row=2, min_col=2, max_col=6):
    for cell in row:
        cell.value = None

wb.active = new_ws
wb.save("Workout.xlsx")
wb.close()